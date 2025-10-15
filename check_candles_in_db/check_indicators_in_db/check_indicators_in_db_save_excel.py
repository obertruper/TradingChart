#!/usr/bin/env python3
"""
Optimized script to check indicators data availability in database by day.
Creates Excel report with separate sheets for each symbol/timeframe/indicator combination.
"""

import os
import sys
import psycopg2
from psycopg2.extras import DictCursor
from datetime import datetime, timedelta
import pandas as pd
from typing import Dict, List
import logging
from dotenv import load_dotenv

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', '82.25.115.144'),
    'port': os.getenv('DB_PORT', 5432),
    'database': os.getenv('DB_NAME', 'trading_data'),
    'user': os.getenv('DB_READER_USER', 'trading_reader'),
    'password': os.getenv('DB_READER_PASSWORD')
}

# Hardcoded list of symbols to check
SYMBOLS = ['BTCUSDT']

# Timeframes to check
TIMEFRAMES = ['1m', '15m', '1h']

# Indicators to check (column names in DB)
INDICATORS = {
    'sma': ['sma_10', 'sma_30', 'sma_50', 'sma_100', 'sma_200'],
    'ema': ['ema_9', 'ema_12', 'ema_21', 'ema_26', 'ema_50', 'ema_100', 'ema_200'],
    'rsi': ['rsi_7', 'rsi_9', 'rsi_14', 'rsi_21', 'rsi_25'],
    'vma': ['vma_10', 'vma_20', 'vma_50', 'vma_100', 'vma_200'],  # Volume Moving Average
    'fear_greed': ['fear_and_greed_index_alternative'],  # Fear & Greed Index from Alternative.me (only for BTCUSDT)
    'coinmarketcap_fear_greed': ['fear_and_greed_index_coinmarketcap']  # Fear & Greed Index from CoinMarketCap (only for BTCUSDT)
}

# Date range for checking (last 30 days by default)
END_DATE = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
START_DATE = END_DATE - timedelta(days=30)

# Analyze full data period from database
FULL_DAYS_DATA = True  # Set to True to analyze all available data


class OptimizedIndicatorChecker:
    def __init__(self):
        self.conn = None
        self.cursor = None

    def connect(self):
        """Connect to database"""
        try:
            self.conn = psycopg2.connect(**DB_CONFIG)
            self.cursor = self.conn.cursor(cursor_factory=DictCursor)
            logger.info("Connected to database")
        except Exception as e:
            logger.error(f"Failed to connect to database: {e}")
            raise

    def disconnect(self):
        """Disconnect from database"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
        logger.info("Disconnected from database")

    def get_data_period(self) -> tuple:
        """Get the earliest and latest dates from indicators tables"""

        min_date = None
        max_date = None

        for timeframe in TIMEFRAMES:
            table_name = f"indicators_bybit_futures_{timeframe}"

            query = f"""
                SELECT
                    MIN(DATE(timestamp)) as min_date,
                    MAX(DATE(timestamp)) as max_date
                FROM {table_name}
                WHERE symbol = %s
            """

            try:
                for symbol in SYMBOLS:
                    self.cursor.execute(query, (symbol,))
                    result = self.cursor.fetchone()

                    if result and result['min_date']:
                        if min_date is None or result['min_date'] < min_date:
                            min_date = result['min_date']
                        if max_date is None or result['max_date'] > max_date:
                            max_date = result['max_date']
            except Exception as e:
                logger.warning(f"Error getting data period from {table_name}: {e}")

        return min_date, max_date

    def fetch_all_data_for_timeframe(self, symbol: str, timeframe: str) -> pd.DataFrame:
        """Fetch all data for a symbol and timeframe"""

        table_name = f"indicators_bybit_futures_{timeframe}"

        # Get all indicator columns
        all_columns = []
        for indicator_cols in INDICATORS.values():
            all_columns.extend(indicator_cols)

        # Build query parts for different indicators
        query_parts = []
        for col in all_columns:
            if col == 'fear_and_greed_index_alternative':
                # Special handling for Fear & Greed from Alternative.me - check if value exists and get classification
                query_parts.append(f"""
                    COUNT(DISTINCT {col}) as {col}_count,
                    MAX({col}) as {col}_max,
                    MAX(fear_and_greed_index_classification_alternative) as fear_and_greed_index_classification_alternative
                """)
            elif col == 'fear_and_greed_index_coinmarketcap':
                # Special handling for Fear & Greed from CoinMarketCap - check if value exists and get classification
                query_parts.append(f"""
                    COUNT(DISTINCT {col}) as {col}_count,
                    MAX({col}) as {col}_max,
                    MAX(fear_and_greed_index_coinmarketcap_classification) as fear_and_greed_index_coinmarketcap_classification
                """)
            else:
                # Regular indicators
                query_parts.append(f"COUNT({col}) as {col}_count, MAX({col}) as {col}_max")

        query = f"""
            SELECT
                DATE(timestamp) as date,
                {', '.join(query_parts)}
            FROM {table_name}
            WHERE symbol = %s
              AND timestamp >= %s
              AND timestamp < %s + interval '1 day'
            GROUP BY DATE(timestamp)
            ORDER BY DATE(timestamp)
        """

        try:
            self.cursor.execute(query, (symbol, START_DATE, END_DATE))
            results = self.cursor.fetchall()

            if results:
                # Extract column names from cursor description
                columns = [desc[0] for desc in self.cursor.description]
                df = pd.DataFrame(results, columns=columns)
                return df
            else:
                return pd.DataFrame()

        except Exception as e:
            logger.error(f"Error fetching data for {symbol} {timeframe}: {e}")
            return pd.DataFrame()

    def generate_sheets(self) -> Dict[str, pd.DataFrame]:
        """Generate individual sheets for each symbol/timeframe/indicator combination"""

        sheets = {}

        # Generate full date range
        date_range = pd.date_range(START_DATE, END_DATE, freq='D')

        # Expected counts per timeframe
        expected_counts = {
            '1m': 1440,
            '15m': 96,
            '1h': 24
        }

        total_combinations = len(SYMBOLS) * len(TIMEFRAMES)
        combination_count = 0

        logger.info(f"Starting data check for {len(SYMBOLS)} symbols, {len(TIMEFRAMES)} timeframes")

        for symbol in SYMBOLS:
            logger.info(f"Processing symbol: {symbol}")

            for timeframe in TIMEFRAMES:
                combination_count += 1
                logger.info(f"  Fetching data for {timeframe} ({combination_count}/{total_combinations})")

                # Fetch all data for this timeframe
                data_df = self.fetch_all_data_for_timeframe(symbol, timeframe)
                expected_count = expected_counts[timeframe]

                # Process each indicator type
                for indicator_type, indicator_columns in INDICATORS.items():
                    sheet_name = f"{symbol}_{timeframe}_{indicator_type}"
                    sheet_data = []

                    for date in date_range:
                        date_str = date.strftime('%Y-%m-%d')

                        # Check if we have data for this date
                        if not data_df.empty and 'date' in data_df.columns:
                            date_data = data_df[data_df['date'] == date.date()]
                        else:
                            date_data = pd.DataFrame()

                        row = {
                            'date': date_str,
                            'day_of_week': date.strftime('%A')[:3],  # Mon, Tue, etc.
                        }

                        if date_data.empty:
                            # No data for this date
                            row['status'] = 'NO_DATA'
                            row['completeness_%'] = 0
                            row['total_candles'] = 0

                            for col in indicator_columns:
                                row[col] = 0

                        else:
                            # Special handling for Fear & Greed Index from Alternative.me
                            if indicator_type == 'fear_greed':
                                # Fear & Greed only applies to BTCUSDT
                                if symbol == 'BTCUSDT':
                                    count_col = "fear_and_greed_index_alternative_count"
                                    max_col = "fear_and_greed_index_alternative_max"
                                    class_col = "fear_and_greed_index_classification_alternative"

                                    if count_col in date_data.columns:
                                        count_val = date_data.iloc[0][count_col]
                                        if pd.notna(count_val) and count_val > 0:
                                            row['fear_and_greed_index_alternative'] = date_data.iloc[0][max_col] if max_col in date_data.columns else 0
                                            row['classification'] = date_data.iloc[0][class_col] if class_col in date_data.columns else 'N/A'
                                            row['completeness_%'] = 100.0  # If we have data, it's complete
                                            row['status'] = 'COMPLETE'
                                            row['total_candles'] = expected_count  # All candles have the same value
                                        else:
                                            row['fear_and_greed_index_alternative'] = 0
                                            row['classification'] = 'NO_DATA'
                                            row['completeness_%'] = 0
                                            row['status'] = 'NO_DATA'
                                            row['total_candles'] = 0
                                    else:
                                        row['fear_and_greed_index_alternative'] = 0
                                        row['classification'] = 'NO_DATA'
                                        row['completeness_%'] = 0
                                        row['status'] = 'NO_DATA'
                                        row['total_candles'] = 0
                                else:
                                    # Fear & Greed is not applicable for non-BTCUSDT symbols
                                    row['fear_and_greed_index_alternative'] = 'N/A'
                                    row['classification'] = 'N/A (BTC only)'
                                    row['completeness_%'] = 0
                                    row['status'] = 'N/A'
                                    row['total_candles'] = 0
                            # Special handling for Fear & Greed Index from CoinMarketCap
                            elif indicator_type == 'coinmarketcap_fear_greed':
                                # CoinMarketCap Fear & Greed only applies to BTCUSDT
                                if symbol == 'BTCUSDT':
                                    count_col = "fear_and_greed_index_coinmarketcap_count"
                                    max_col = "fear_and_greed_index_coinmarketcap_max"
                                    class_col = "fear_and_greed_index_coinmarketcap_classification"

                                    if count_col in date_data.columns:
                                        count_val = date_data.iloc[0][count_col]
                                        if pd.notna(count_val) and count_val > 0:
                                            row['fear_and_greed_index_coinmarketcap'] = date_data.iloc[0][max_col] if max_col in date_data.columns else 0
                                            row['classification'] = date_data.iloc[0][class_col] if class_col in date_data.columns else 'N/A'
                                            row['completeness_%'] = 100.0  # If we have data, it's complete
                                            row['status'] = 'COMPLETE'
                                            row['total_candles'] = expected_count  # All candles have the same value
                                        else:
                                            row['fear_and_greed_index_coinmarketcap'] = 0
                                            row['classification'] = 'NO_DATA'
                                            row['completeness_%'] = 0
                                            row['status'] = 'NO_DATA'
                                            row['total_candles'] = 0
                                    else:
                                        row['fear_and_greed_index_coinmarketcap'] = 0
                                        row['classification'] = 'NO_DATA'
                                        row['completeness_%'] = 0
                                        row['status'] = 'NO_DATA'
                                        row['total_candles'] = 0
                                else:
                                    # CoinMarketCap Fear & Greed is not applicable for non-BTCUSDT symbols
                                    row['fear_and_greed_index_coinmarketcap'] = 'N/A'
                                    row['classification'] = 'N/A (BTC only)'
                                    row['completeness_%'] = 0
                                    row['status'] = 'N/A'
                                    row['total_candles'] = 0
                            else:
                                # Regular indicators (SMA, EMA, RSI)
                                total_values = 0
                                max_possible = 0

                                for col in indicator_columns:
                                    count_col = f"{col}_count"
                                    if count_col in date_data.columns:
                                        count_val = date_data.iloc[0][count_col]
                                        row[col] = count_val if pd.notna(count_val) else 0
                                        total_values += row[col]
                                    else:
                                        row[col] = 0

                                    max_possible += expected_count

                                # Calculate completeness
                                completeness = (total_values / max_possible * 100) if max_possible > 0 else 0
                                row['completeness_%'] = round(completeness, 2)

                                # Determine status
                                if completeness >= 95:
                                    row['status'] = 'COMPLETE'
                                elif completeness >= 50:
                                    row['status'] = 'PARTIAL'
                                elif completeness > 0:
                                    row['status'] = 'INSUFFICIENT'
                                else:
                                    row['status'] = 'NO_INDICATORS'

                                # Calculate total candles (average across indicators)
                                indicator_values = [row[col] for col in indicator_columns if row[col] > 0]
                                row['total_candles'] = max(indicator_values) if indicator_values else 0

                        # Add expected count
                        row['expected_candles'] = expected_count

                        sheet_data.append(row)

                    # Create DataFrame for this sheet
                    df = pd.DataFrame(sheet_data)

                    # Reorder columns for better readability
                    base_cols = ['date', 'day_of_week', 'status', 'completeness_%', 'total_candles', 'expected_candles']

                    # Special columns for Fear & Greed from Alternative.me
                    if indicator_type == 'fear_greed':
                        if 'fear_and_greed_index_alternative' in df.columns:
                            extra_cols = ['fear_and_greed_index_alternative', 'classification']
                        else:
                            extra_cols = []
                    # Special columns for Fear & Greed from CoinMarketCap
                    elif indicator_type == 'coinmarketcap_fear_greed':
                        if 'fear_and_greed_index_coinmarketcap' in df.columns:
                            extra_cols = ['fear_and_greed_index_coinmarketcap', 'classification']
                        else:
                            extra_cols = []
                    else:
                        extra_cols = [col for col in indicator_columns if col in df.columns]

                    ordered_cols = base_cols + extra_cols
                    # Only include columns that exist
                    ordered_cols = [col for col in ordered_cols if col in df.columns]

                    df = df[ordered_cols]
                    sheets[sheet_name] = df

                    # Log progress
                    complete_days = len(df[df['status'] == 'COMPLETE'])
                    logger.info(f"    {sheet_name}: {complete_days}/{len(df)} complete days")

        logger.info("Data check completed")
        return sheets

    def create_summary_sheet(self, sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Create summary statistics sheet"""

        summary_data = []

        for sheet_name, df in sheets.items():
            parts = sheet_name.split('_')
            symbol = parts[0]
            timeframe = parts[1]
            indicator = parts[2].upper()

            # Calculate stats differently for N/A status (non-BTC Fear & Greed)
            if 'N/A' in df['status'].values:
                # Skip N/A rows in calculations for Fear & Greed on non-BTC symbols
                valid_df = df[df['status'] != 'N/A']
                if len(valid_df) > 0:
                    avg_completeness = round(valid_df['completeness_%'].mean(), 2)
                    min_completeness = round(valid_df['completeness_%'].min(), 2)
                    max_completeness = round(valid_df['completeness_%'].max(), 2)
                    data_coverage = round(len(valid_df[valid_df['status'] != 'NO_DATA']) / len(valid_df) * 100, 2) if len(valid_df) > 0 else 0
                else:
                    avg_completeness = 0
                    min_completeness = 0
                    max_completeness = 0
                    data_coverage = 0

                summary_data.append({
                    'symbol': symbol,
                    'timeframe': timeframe,
                    'indicator': indicator,
                    'total_days': len(df),
                    'complete_days': len(df[df['status'] == 'COMPLETE']),
                    'partial_days': len(df[df['status'] == 'PARTIAL']),
                    'insufficient_days': len(df[df['status'] == 'INSUFFICIENT']),
                    'no_data_days': len(df[df['status'].isin(['NO_DATA', 'NO_INDICATORS'])]),
                    'n/a_days': len(df[df['status'] == 'N/A']),
                    'avg_completeness_%': avg_completeness,
                    'min_completeness_%': min_completeness,
                    'max_completeness_%': max_completeness,
                    'data_coverage_%': data_coverage
                })
            else:
                summary_data.append({
                    'symbol': symbol,
                    'timeframe': timeframe,
                    'indicator': indicator,
                    'total_days': len(df),
                    'complete_days': len(df[df['status'] == 'COMPLETE']),
                    'partial_days': len(df[df['status'] == 'PARTIAL']),
                    'insufficient_days': len(df[df['status'] == 'INSUFFICIENT']),
                    'no_data_days': len(df[df['status'].isin(['NO_DATA', 'NO_INDICATORS'])]),
                    'n/a_days': 0,
                    'avg_completeness_%': round(df['completeness_%'].mean(), 2),
                    'min_completeness_%': round(df['completeness_%'].min(), 2),
                    'max_completeness_%': round(df['completeness_%'].max(), 2),
                    'data_coverage_%': round(len(df[df['status'] != 'NO_DATA']) / len(df) * 100, 2)
                })

        return pd.DataFrame(summary_data)


def save_excel_report(filename: str, sheets: Dict[str, pd.DataFrame], summary_df: pd.DataFrame):
    """Save all sheets to a single Excel file"""

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.formatting.rule import CellIsRule

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Save summary sheet first
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Save individual sheets
            for sheet_name, df in sheets.items():
                # Excel sheet name limit is 31 characters
                truncated_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                df.to_excel(writer, sheet_name=truncated_name, index=False)

            # Get the workbook and apply formatting
            workbook = writer.book

            # Format summary sheet
            ws_summary = workbook['Summary']
            for row in ws_summary['A1:J1']:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)

            # Format individual sheets
            for sheet_name in workbook.sheetnames[1:]:  # Skip Summary
                ws = workbook[sheet_name]

                # Header formatting
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)

                # Conditional formatting for status column
                status_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'status':
                        status_col = openpyxl.utils.get_column_letter(idx)
                        break

                if status_col:
                    # Color coding for status
                    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{status_col}{row}']
                        if cell.value == 'COMPLETE':
                            cell.fill = green_fill
                        elif cell.value == 'PARTIAL':
                            cell.fill = yellow_fill
                        elif cell.value in ['NO_DATA', 'NO_INDICATORS', 'INSUFFICIENT']:
                            cell.fill = red_fill
                        elif cell.value == 'N/A':
                            cell.fill = gray_fill

                # Special formatting for Fear & Greed sheets (both Alternative.me and CoinMarketCap)
                if 'fear_greed' in sheet_name.lower():
                    # Find Fear & Greed index column
                    fg_index_col = None
                    classification_col = None
                    for idx, cell in enumerate(ws[1], 1):
                        if cell.value in ('fear_and_greed_index_alternative', 'fear_and_greed_index_coinmarketcap'):
                            fg_index_col = openpyxl.utils.get_column_letter(idx)
                        elif cell.value == 'classification':
                            classification_col = openpyxl.utils.get_column_letter(idx)

                    if fg_index_col:
                        # Color coding for Fear & Greed values
                        extreme_fear_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # Dark red
                        fear_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Red
                        neutral_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')  # Yellow
                        greed_fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')  # Light green
                        extreme_greed_fill = PatternFill(start_color='00A651', end_color='00A651', fill_type='solid')  # Green

                        for row in range(2, ws.max_row + 1):
                            cell = ws[f'{fg_index_col}{row}']
                            if isinstance(cell.value, (int, float)):
                                value = int(cell.value)
                                if 0 <= value <= 25:
                                    cell.fill = extreme_fear_fill
                                    cell.font = Font(color='FFFFFF', bold=True)
                                elif 26 <= value <= 45:
                                    cell.fill = fear_fill
                                    cell.font = Font(color='FFFFFF')
                                elif 46 <= value <= 55:
                                    cell.fill = neutral_fill
                                elif 56 <= value <= 75:
                                    cell.fill = greed_fill
                                elif 76 <= value <= 100:
                                    cell.fill = extreme_greed_fill
                                    cell.font = Font(color='FFFFFF', bold=True)

                    if classification_col:
                        # Apply same colors to classification column
                        for row in range(2, ws.max_row + 1):
                            cell = ws[f'{classification_col}{row}']
                            if cell.value == 'Extreme Fear':
                                cell.fill = extreme_fear_fill
                                cell.font = Font(color='FFFFFF', bold=True)
                            elif cell.value == 'Fear':
                                cell.fill = fear_fill
                                cell.font = Font(color='FFFFFF')
                            elif cell.value == 'Neutral':
                                cell.fill = neutral_fill
                            elif cell.value == 'Greed':
                                cell.fill = greed_fill
                            elif cell.value == 'Extreme Greed':
                                cell.fill = extreme_greed_fill
                                cell.font = Font(color='FFFFFF', bold=True)

                # Auto-adjust column widths
                for column_cells in ws.columns:
                    length = max(len(str(cell.value or '')) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        logger.info(f"Excel file saved successfully: {filename}")
        return True

    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return False


def main():
    """Main execution function"""

    global START_DATE, END_DATE

    logger.info("=" * 80)
    logger.info("Optimized Indicator Data Checker with Excel Export")

    # Create results directory
    results_dir = os.path.join(os.path.dirname(__file__), 'results')
    os.makedirs(results_dir, exist_ok=True)

    checker = OptimizedIndicatorChecker()

    try:
        # Connect to database
        checker.connect()

        # Determine date range based on FULL_DAYS_DATA setting
        if FULL_DAYS_DATA:
            logger.info("FULL_DAYS_DATA is True - analyzing all available data")
            logger.info("Fetching data period from database...")

            db_min_date, db_max_date = checker.get_data_period()

            if db_min_date and db_max_date:
                START_DATE = datetime.combine(db_min_date, datetime.min.time())
                END_DATE = datetime.combine(db_max_date, datetime.min.time())

                total_days = (db_max_date - db_min_date).days + 1
                logger.info(f"Database contains data from {db_min_date} to {db_max_date}")
                logger.info(f"Total period: {total_days} days")

                if total_days > 365:
                    logger.warning(f"⚠️  Large dataset detected: {total_days} days")
                    logger.warning("This analysis may take several minutes...")
            else:
                logger.error("Could not determine data period from database")
                logger.info("Falling back to last 30 days")
                END_DATE = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                START_DATE = END_DATE - timedelta(days=30)
        else:
            logger.info("FULL_DAYS_DATA is False - analyzing last 30 days")

        logger.info(f"Date range: {START_DATE.date()} to {END_DATE.date()}")
        logger.info(f"Symbols: {', '.join(SYMBOLS)}")
        logger.info(f"Timeframes: {', '.join(TIMEFRAMES)}")
        logger.info(f"Indicators: {', '.join([i.upper() for i in INDICATORS.keys()])}")
        logger.info("=" * 80)

        # Generate all sheets
        logger.info("Generating indicator analysis sheets...")
        sheets = checker.generate_sheets()

        # Create summary
        logger.info("Creating summary sheet...")
        summary_df = checker.create_summary_sheet(sheets)

        # Save to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = os.path.join(results_dir, f"indicators_analysis_{timestamp}.xlsx")

        logger.info(f"Saving to Excel file: {excel_filename}")
        if save_excel_report(excel_filename, sheets, summary_df):
            logger.info(f"✅ Report saved successfully: {excel_filename}")
        else:
            logger.error("Failed to save Excel report")

        # Print summary
        logger.info("\n" + "=" * 80)
        logger.info("ANALYSIS SUMMARY")
        logger.info("=" * 80)

        for _, row in summary_df.iterrows():
            logger.info(f"\n{row['symbol']} - {row['indicator']} - {row['timeframe']}:")
            logger.info(f"  Complete: {row['complete_days']}/{row['total_days']} days")
            logger.info(f"  Coverage: {row['data_coverage_%']}%")
            logger.info(f"  Avg Completeness: {row['avg_completeness_%']}%")

    except Exception as e:
        logger.error(f"Error during execution: {e}")
        raise

    finally:
        checker.disconnect()

    logger.info("\n" + "=" * 80)
    logger.info("Analysis complete!")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()